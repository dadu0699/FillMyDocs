from io import BytesIO
from os import path

from jinja2 import Template
from openpyxl import load_workbook

from app.services.pdf_service import convert_to_pdf

TEMPLATE_DIR = "app/templates/"
template_cache = {}


async def render_xlsx_template(template_name: str, context: dict, get_pdf: bool = False):
    """
    Procesa un archivo .xlsx utilizando el contexto proporcionado.
    Si get_pdf es True, convierte el archivo a PDF antes de devolverlo.
    Devuelve un BytesIO con el contenido del archivo (xlsx o pdf) y un mensaje de error o None.
    """
    template_path = path.join(TEMPLATE_DIR, template_name)

    if template_name not in template_cache:
        try:
            template_cache[template_name] = template_path
        except FileNotFoundError:
            return None, f"Template not found: {template_name}"

    try:
        workbook = load_workbook(template_cache[template_name])
        worksheet = workbook.active

        if worksheet is None:
            return None, "No active worksheet found in template."

        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    template = Template(cell.value)
                    cell.value = template.render(context)

        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        if get_pdf:
            pdf_stream, error = await convert_to_pdf(file_stream, ".xlsx")
            if error:
                return None, error
            return pdf_stream, None

        return file_stream, None

    except Exception as exception:
        return None, str(exception)
